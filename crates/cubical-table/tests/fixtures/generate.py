#!/usr/bin/env python3
"""Regenerates the .xlsx fixtures used by cubical-table's Excel tests.

Run: python3 crates/cubical-table/tests/fixtures/generate.py

The .xlsx files it writes are committed, so the test suite never needs Python
or openpyxl at run time. Written with openpyxl 3.1.5.

Two things openpyxl cannot express directly are patched into the saved zip:

  * a cached formula result. openpyxl never evaluates a formula, so it writes
    <f>SUM(...)</f> with no <v>. Excel always stores the last computed value,
    and calamine reads that cached <v> — so the fixture is patched to carry one,
    which is the case the reader actually meets in the wild.
  * a padded used-range. A trailing row and column hold a single space, so the
    used range covers cells that carry nothing a reader can use — the residue a
    spreadsheet app routinely leaves behind, and what the trimming logic exists
    to absorb. Styling alone does not do it: calamine drops a styled cell that
    has no value, so the fixture would exercise nothing.
"""

import re
import shutil
import zipfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

HERE = Path(__file__).resolve().parent
WORKBOOK = HERE / "workbook.xlsx"


def build() -> None:
    wb = Workbook()

    q1 = wb.active
    q1.title = "Q1"
    q1.append(["name", "qty", "ok", "due", "code"])
    q1.append(["Alpha", 2, True, date(2024, 1, 15), "007"])
    q1.append(["Beta", 1.5, False, datetime(2024, 3, 2, 13, 45, 30), None])
    q1.append(["Total", "=SUM(B2:B3)", None, None, "x"])
    q1["F1"] = " "
    q1["F1"].font = Font(bold=True)
    q1["A5"] = " "
    q1["A5"].font = Font(bold=True)

    q2 = wb.create_sheet("Q2")
    q2.append(["city", "population"])
    q2.append(["Oslo", 709037])

    wb.create_sheet("Q3")

    wb.save(WORKBOOK)


def patch_zip() -> None:
    source = WORKBOOK.with_suffix(".tmp")
    shutil.move(WORKBOOK, source)
    with zipfile.ZipFile(source) as old:
        names = old.namelist()
        blobs = {name: old.read(name) for name in names}
    sheet = "xl/worksheets/sheet1.xml"
    xml = blobs[sheet].decode("utf-8")
    patched = re.sub(
        r"<f>SUM\(B2:B3\)</f>(<v></v>)?",
        "<f>SUM(B2:B3)</f><v>3.5</v>",
        xml,
        count=1,
    )
    if patched == xml:
        raise SystemExit(f"{sheet}: no SUM(B2:B3) formula to give a cached value")
    blobs[sheet] = patched.encode("utf-8")
    with zipfile.ZipFile(WORKBOOK, "w", zipfile.ZIP_DEFLATED) as new:
        for name in names:
            new.writestr(name, blobs[name])
    source.unlink()


if __name__ == "__main__":
    build()
    patch_zip()
    print(f"wrote {WORKBOOK.relative_to(HERE.parents[2])}")
